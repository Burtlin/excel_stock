"""
股票數據處理主程式
整合營收、財務報表、EPS數據處理
"""
import logging
import os
import sys
from datetime import datetime

import pandas as pd
from FinMind.data import DataLoader
from openpyxl.styles import Border, Side

# 導入配置
from config import BASE_DIR

# 導入模組
from modules.logger import setup_logging, clean_old_logs
from modules.utils import process_info_data, format_percentage_columns
from modules.revenue import process_revenue_data, get_previous_three_months, create_revenue_overview
from modules.financial import process_financial_data, process_eps_data, create_financial_overview


def is_file_open(file_path):
    """檢查檔案是否被其他程式（如Excel）開啟"""
    try:
        # 嘗試以獨佔寫入模式開啟檔案
        with open(file_path, 'a') as f:
            pass
        return False
    except IOError:
        return True


def process_stock(input_file='target.xlsx', output_file=None, revenue_sheet='月營收', financial_sheet='綜合損益表', eps_sheet='EPS', overview_sheet='營收總覽', financial_overview_sheet='財務總覽'):
    """處理股票數據，營收、財務和 EPS 數據分別輸出到不同的 sheet"""
    # 初始化 logging
    setup_logging()
    clean_old_logs(days=7)
    
    logging.info("="*60)
    logging.info("開始處理股票數據")
    logging.info(f"輸入檔案: {input_file}")
    
    api = DataLoader()
    # api.login_by_token(api_token='token')
    # api.login(user_id='user_id', password='password')

    # 讀取第一個 sheet 取得股票代號
    df_base = pd.read_excel(input_file, sheet_name=0)
    df_base = df_base[['代號']].astype(int)
    
    # 創建三個 DataFrame：營收、綜合損益表、EPS
    df_revenue = df_base.copy()
    df_financial = df_base.copy()
    df_eps = df_base.copy()
    
    # 加入名稱欄位
    df_revenue = process_info_data(api, df_revenue)
    df_financial = process_info_data(api, df_financial)
    df_eps = process_info_data(api, df_eps)
    
    # 使用輔助函數計算時間
    (last_month_year, last_month), (previous_month_year, previous_month), (previous_month_year2, previous_month2) = get_previous_three_months()
    
    # 計算去年同期
    yoy_year = last_month_year - 1
    
    total = len(df_base)
    for idx, row in df_base.iterrows():
        stock_id = row["代號"]
        progress_msg = f"[{idx+1}/{total}] 處理中: {stock_id}"
        logging.info(progress_msg)
        print(f"  {progress_msg}")  # 同時輸出到控制台
        
        # 處理營收數據（寫入 df_revenue）
        process_revenue_data(api, df_revenue, idx, stock_id, last_month_year, last_month, previous_month_year, previous_month, previous_month_year2, previous_month2, yoy_year)
        
        # 處理綜合損益表數據（寫入 df_financial）
        try:
            process_financial_data(api, df_financial, idx, stock_id)
        except Exception as e:
            logging.error(f"  錯誤: {stock_id} 財務數據處理失敗 - {str(e)}")
        
        # 處理 EPS 數據（寫入 df_eps）
        try:
            process_eps_data(api, df_eps, idx, stock_id, last_month_year)
        except Exception as e:
            logging.error(f"  錯誤: {stock_id} EPS 數據處理失敗 - {str(e)}")
    
    logging.info("\n處理完成！")
    logging.info(f"\n營收數據:\n{df_revenue.head()}")
    logging.info(f"\n綜合損益表數據:\n{df_financial.head()}")
    logging.info(f"\nEPS數據:\n{df_eps.head()}")
    
    # 生成營收總覽（橫向格式：月份為列，股票為欄）
    logging.info("\n生成營收總覽...")
    print("\n✓ 月營收處理完成")
    print("  正在生成營收總覽...")
    df_overview = create_revenue_overview(api, df_base)
    logging.info(f"\n營收總覽數據:\n{df_overview.head()}")
    
    # 生成財務總覽（橫向格式：8季度）
    logging.info("\n生成財務總覽...")
    print("✓ 營收總覽完成")
    print("  正在生成財務總覽...")
    df_financial_overview = create_financial_overview(api, df_base)
    logging.info(f"\n財務總覽數據:\n{df_financial_overview.head()}")
    print("✓ 財務總覽完成")
    
    # 決定輸出檔案
    if output_file is None:
        output_file = input_file
    
    # 使用 openpyxl 保留原檔案的其他 sheet 和格式
    try:
        # 使用 ExcelWriter 將五個 DataFrame 分別寫入不同 sheet
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df_revenue.to_excel(writer, sheet_name=revenue_sheet, index=False)
            df_financial.to_excel(writer, sheet_name=financial_sheet, index=False)
            df_eps.to_excel(writer, sheet_name=eps_sheet, index=False)
            df_overview.to_excel(writer, sheet_name=overview_sheet, index=False)
            df_financial_overview.to_excel(writer, sheet_name=financial_overview_sheet, index=False)
            
            # 格式化營收總覽的百分比欄位（MoM和YoY行）並添加邊框
            ws_overview = writer.sheets[overview_sheet]
            
            # 定義上邊框樣式
            top_border = Border(top=Side(style='thin', color='808080'))
            
            for row_idx in range(2, ws_overview.max_row + 1):
                cell_code = ws_overview.cell(row=row_idx, column=1)
                cell_name = ws_overview.cell(row=row_idx, column=2)
                
                # 檢查是否為股票開始行（代號欄位有值且名稱欄位不是 MoM/YoY）
                if cell_code.value and cell_name.value not in ['MoM(%)', 'YoY(%)']:
                    # 為該行所有欄位添加上邊框
                    for col_idx in range(1, ws_overview.max_column + 1):
                        cell = ws_overview.cell(row=row_idx, column=col_idx)
                        cell.border = top_border
                
                # 檢查名稱欄位是否為 'MoM(%)' 或 'YoY(%)'
                if cell_name.value in ['MoM(%)', 'YoY(%)']:
                    # 格式化該行的所有百分比值（從第3欄開始，跳過代號和名稱）
                    for col_idx in range(3, ws_overview.max_column + 1):
                        cell = ws_overview.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                            cell.value = cell.value / 100
            
            # 格式化財務總覽的百分比欄位和邊框
            ws_financial_overview = writer.sheets[financial_overview_sheet]
            
            for row_idx in range(2, ws_financial_overview.max_row + 1):
                cell_code = ws_financial_overview.cell(row=row_idx, column=1)
                cell_name = ws_financial_overview.cell(row=row_idx, column=2)
                
                # 檢查是否為股票開始行（代號欄位有值）
                if cell_code.value and cell_name.value not in ['毛利率', '營益率', '淨利率', 'EPS']:
                    # 為該行所有欄位添加上邊框
                    for col_idx in range(1, ws_financial_overview.max_column + 1):
                        cell = ws_financial_overview.cell(row=row_idx, column=col_idx)
                        cell.border = top_border
                
                # 檢查名稱欄位是否為百分比行（毛利率、營益率、淨利率）
                if cell_name.value in ['毛利率', '營益率', '淨利率']:
                    # 格式化該行的所有百分比值（從第3欄開始，跳過代號和名稱）
                    for col_idx in range(3, ws_financial_overview.max_column + 1):
                        cell = ws_financial_overview.cell(row=row_idx, column=col_idx)
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                            cell.value = cell.value / 100
            
            # 格式化百分比欄位
            format_percentage_columns(writer.sheets[revenue_sheet], df_revenue)
            format_percentage_columns(writer.sheets[financial_sheet], df_financial)
        
        logging.info(f"\n已更新並儲存至: {output_file}")
        logging.info(f"  - 營收總覽: {overview_sheet}")
        logging.info(f"  - 財務總覽: {financial_overview_sheet}")
        logging.info(f"  - 營收數據: {revenue_sheet}")
        logging.info(f"  - 綜合損益表: {financial_sheet}")
        logging.info(f"  - EPS數據: {eps_sheet}")
    except Exception as e:
        logging.error(f"\n儲存檔案時發生錯誤: {str(e)}")
    
    logging.info("處理完成")
    logging.info("="*60 + "\n")
    
    return df_revenue, df_financial, df_eps


def main():
    """主程式進入點，增加錯誤處理"""
    import time
    
    print("\n" + "=" * 60)
    print("            股票數據處理系統")
    print("=" * 60)
    
    try:
        # 自動使用 target.xlsx
        input_file = os.path.join(BASE_DIR, 'target.xlsx')
        
        # 檢查檔案是否存在
        if not os.path.exists(input_file):
            print(f"\n❌ 錯誤: 找不到 target.xlsx")
            print(f"   請將 target.xlsx 放在程式目錄中")
            input("\n按 Enter 鍵離開...")
            return 1
        
        # 檢查檔案是否被開啟
        if is_file_open(input_file):
            print(f"\n⚠ 警告: target.xlsx 正在被使用中（可能被 Excel 開啟）")
            print(f"   請關閉 Excel 後繼續...\n")
            
            # 等待使用者關閉檔案
            while is_file_open(input_file):
                response = input("關閉檔案後按 Enter 繼續，或輸入 'q' 離開: ")
                if response.lower() == 'q':
                    print("\n已取消執行")
                    return 0
                if not is_file_open(input_file):
                    print("\n✓ 檔案已可使用\n")
                    break
                else:
                    print("⚠ 檔案仍在使用中，請確認已關閉 Excel\n")
        
        # 顯示檔案資訊
        df_check = pd.read_excel(input_file, sheet_name=0)
        stock_count = len(df_check)
        print(f"✓ 找到 target.xlsx ({stock_count} 支股票)")
        print(f"\n開始處理...\n")
        
        start_time = time.time()
        
        # 執行處理
        process_stock(input_file=input_file, output_file=input_file)
        
        elapsed = time.time() - start_time
        
        # 成功訊息
        print("\n" + "=" * 60)
        print("✓ 處理完成！")
        print(f"  - 成功處理: {stock_count} 支股票")
        print(f"  - 耗時: {elapsed:.1f} 秒")
        print(f"  - 輸出檔案: {input_file}")
        print("=" * 60)
        
        # 自動開啟 Excel
        try:
            print("\n正在開啟 Excel...")
            os.startfile(input_file)
        except Exception as e:
            print(f"⚠ 無法自動開啟 Excel: {str(e)}")
        
        return 0
        
    except Exception as e:
        print(f"\n❌ 程式執行失敗: {str(e)}")
        import traceback
        traceback.print_exc()
        input("\n按 Enter 鍵離開...")
        return 1


if __name__ == '__main__':
    sys.exit(main())
